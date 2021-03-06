import pandas as pd
import os
from pycaret.regression import *
import pmdarima as pm
from pmdarima.arima import ARIMA
from openpyxl import load_workbook, Workbook

import matplotlib.pyplot as plt
import seaborn as sns

def train_test_combine_split(df, train_end_dt, test_start_dt, test_end_dt, date_column='Calendar Date'):
    """

    :param df: dataframe
    :param train_end_dt: end training date
    :param test_start_dt: start test date
    :param test_end_dt: end test date
    :param date_column: name of date column
    :return: train, test and comined dataframe
    """

    df[date_column] = pd.to_datetime(df[date_column])

    train_df = df[df[date_column] <= train_end_dt]
    print('Shape of the training dataframe: ', train_df.shape)

    # Test df
    test_df = df[(df[date_column] >= test_start_dt) & (df[date_column] <= test_end_dt)]
    print('Shape of the testing dataframe: ', test_df.shape)

    # Combined dataframe
    comb_df = pd.concat([train_df, test_df])
    print('Shape of the combination dataframe: ', comb_df.shape)

    return train_df, test_df, comb_df

def features_correlation(df, target_col, correlation_threshold=0.5, save_path=False):
    # Run Correlations to target
    corr_df = df.corr()[[target_col]]
    corr_df = corr_df[corr_df[target_col].abs() >= correlation_threshold].sort_values(target_col, ascending=False)

    if save_path:
        file_path = save_path + "/correlations.xlsx"
        if os.path.exists(file_path):
            book = load_workbook(file_path)
        else:
            book = Workbook()
        with pd.ExcelWriter(file_path, engine = 'openpyxl') as writer:
            writer.book = book
            corr_df.to_excel(writer, sheet_name = target_col[0:30], index=True)

    fig, ax = plt.subplots(figsize=(8, 12))
    ax2 = sns.heatmap(corr_df, vmin=-1, vmax=1, annot=True, cmap='BrBG', ax=ax)
    ax.set_title('Features Correlating with '+target_col, fontdict={'fontsize':18}, pad=16)

    corr_df = corr_df.rename_axis('Feature').reset_index()
    corr_feature_cols = corr_df['Feature'].tolist()
    corr_feature_cols.remove(target_col)

    return corr_feature_cols, fig

def run_auto_ml(train_df, test_df, target_col, feature_cols, normal_transform, ml_criteria):
    # Model Definitions
    s = setup(data = train_df,
              test_data = test_df,
              target = target_col,
              fold_strategy = 'timeseries',
              numeric_features = feature_cols,
              fold = 3,
              transform_target = normal_transform,
              feature_selection = True,
              feature_selection_threshold = 0.8,
              remove_multicollinearity = True,
              multicollinearity_threshold = 0.9,
              session_id = 123,
              silent=True)
    best = compare_models(sort = ml_criteria)
    results = pull()
    model_results = results.Model.tolist()
    print(type(best))
    return best

def get_important_features(model, threshold):
    # Get list of features and their values
    features = pd.DataFrame()
    if model == 'xgboost':
        xgb = create_model('xgboost', cross_validation=True, verbose=False)
        feature_important = xgb.get_booster().get_score(importance_type='gain')
        keys = list(feature_important.keys())
        values = list(feature_important.values())
        features = pd.DataFrame(data=values, index=keys, columns=['Variable Importance']).sort_values(by='Variable Importance', ascending=False)
        features = features[features.index.str.contains('Calendar Date') == False]
        features = features.rename_axis('Feature').reset_index()
        features = features[features.index <= threshold-1]
    return features

def run_auto_arima(df, feature_cols, pred_start_dt, forecast_window, ci, alpha = 0.05):
    model_arima = None
    pred_df = pd.DataFrame()
    dti = pd.date_range(pred_start_dt, periods=forecast_window, freq="M")
    dti = dti + pd.offsets.MonthBegin(-1)
    pred_df['Calendar Date'] = dti
    for col in feature_cols:
        print('############################  Running Auto ARIMA for '+col+'   ############################')
        model_arima = pm.auto_arima(df[col],
                              start_p=1,
                              start_q=1,
                              max_p=5,
                              max_q=5,
                              m=12,
                              start_P=0,
                              seasonal=True,
                              d=1,
                              D=1,
                              trace=True,
                              error_action='ignore',  # don't want to know if an order does not work
                              suppress_warnings=True,  # don't want convergence warnings
                              stepwise=True)  # set to stepwise
        # make future predictions
        if ci:
            y_pred, conf_int = model_arima.predict(n_periods=forecast_window, return_conf_int=True, alpha=alpha)
            pred_df[col] = y_pred
            pred_df['Lower CI - {}%'.format(100-alpha*100)], pred_df['Upper CI - {}%'.format(100-alpha*100)] = conf_int.T
        else:
            pred_df[col] = model_arima.predict(n_periods=forecast_window)
        print('############################  End Auto ARIMA for '+col+'   ############################')
        print('')
        print('')
    return pred_df, model_arima

def compute_apes_and_mapes(df, date_col, target_col, feature_cols):
    sort_cols = [date_col, target_col]
    ape_cols = []
    for col in feature_cols:
        ape_col = col+' - APE'
        df[ape_col] = ((df[target_col] - df[col]).abs())/df[target_col]
        sort_cols.append(col)
        sort_cols.append(ape_col)
        ape_cols.append(ape_col)
    df = df[sort_cols]
    non_ape_cols = list(set(sort_cols) - set(ape_cols))
    mape_df = df.groupby([True]*len(df)).mean(numeric_only=True)
    mape_df[non_ape_cols] = np.nan
    df = pd.concat([df,mape_df])
    return df